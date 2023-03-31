# -*- coding: UTF-8 -*-
# @Author  : jierong
# @Time    : 2022/03/15 12:00
# @Function:
import multiprocessing
import sys
import time
import requests
import xlrd
from colorama import Fore, init
from openpyxl import Workbook
init(autoreset=True)


def readUrlFromXlsx(xlsx_file="20220323.xlsx",table_name="Table",col_values=0):
    """
    读取指定xlsx的指定列的数据，并全部返回
    :param xlsx_file: xlsx的文件路径
    :param table_name: 读取的表名
    :param col_values: 读取那一列，从0开始
    :return: 返回该列的所有数据
    """
    excel = xlrd.open_workbook(xlsx_file)
    return excel.sheet_by_name(table_name).col_values(col_values)


def get_result(page):
    rsp = requests.get(page)
    rsp.encoding = "utf-8"
    rsp_time = rsp.elapsed.microseconds
    result = {'resp_code':rsp.status_code,'page':page,'page_size':len(rsp.text)/1024,'rsp_time':rsp_time/1000}
    print(result)
    return result

def main():
    PROCESS_POOL_SIZE = 30

    firstCol = readUrlFromXlsx() # 读取xlsx的页面链接
    total_url_num = len(firstCol)
    firstCol = list(set(firstCol)) # 去重复

    if total_url_num != len(firstCol):
        print(Fore.RED + "警告：存在重复项, 请进行检查")
        print('重复数有：'+ str(total_url_num - len(firstCol)))
        sys.exit(0)

    if total_url_num == 0:
        print(Fore.RED + "警告：未读取到任何页面")
        sys.exit(0)

    print(Fore.RED + "一共需要爬取{}个页面".format(str(total_url_num)))


    pool = multiprocessing.Pool(processes=PROCESS_POOL_SIZE)
    pool_outputs = pool.map(get_result, firstCol)
    pool.close()
    pool.join()
    print("一共扫描的url数量: {} ".format(str(total_url_num)))
    print(Fore.GREEN+"结果集合数量共：{}".format(str(len(pool_outputs))))
    saveUrlToXlsx(pool_outputs) # 保存到xlsx中


def saveUrlToXlsx(results):

    workbook = Workbook()
    result_sheet = workbook.create_sheet(title="result",index=0)
    result_sheet.cell(row=1, column=1,value='请求')
    result_sheet.cell(row=1, column=2,value='响应码')
    result_sheet.cell(row=1, column=3,value='文档大小/KB')
    result_sheet.cell(row=1, column=4,value='响应时间/ms')
    for index, url in enumerate(results):
        # print(url)
        print(url["page"])
        result_sheet.cell(row=index+2, column=1,value=url["page"])
        result_sheet.cell(row=index+2, column=2,value=url["resp_code"])
        result_sheet.cell(row=index+2, column=3,value=url["page_size"])
        result_sheet.cell(row=index+2, column=4,value=url["rsp_time"])

    file_name = "result.xlsx"
    workbook.save(file_name) # 保存到xlsx中
    print("已保存到 {}".format(file_name))
    workbook.close()

if __name__ == '__main__':
    started = time.time()
    main()
    elapsed = time.time() - started
    print("Time elapsed: {:.2f}s".format(elapsed))
    print("Done")
