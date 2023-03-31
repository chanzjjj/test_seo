import requests
import xlrd
from time import time
import threading
from queue import Queue
from openpyxl import Workbook
from dingding import send_dingding


'''
脚本思路：把需要验证403的url放在url_files.xlsx表中，然后把这份表分成两份数据，启用两个线程同时运行这两份数据
'''

excel = xlrd.open_workbook('20220323.xlsx') # 读取excel表
sheet = excel.sheet_by_name("Table") # 读取名为“Table”的sheet
norws = sheet.nrows # 总的url数

url_question = Queue() # 有问题的的url队列，非200和304
url_error = Queue() # 异常的url队列
time_list = [] # 存放每个请求所消耗的时间列表

def run_1():
    for i in range(1,norws//2+1):
        url = sheet.cell(i-1,0).value
        try:
            req = requests.get(url)
            res_code = req.status_code # 请求返回的状态码
            res_time = round(req.elapsed.total_seconds(),4) # 请求时间，保留4位数
            time_list.append(res_time)
            res_size = round(len(req.text)/1024, 2)
            text = "进行到第%s条数据,状态码：%s，响应时间：%ss,url：%s，页面大小：%sKB" %(i, res_code, res_time, url, res_size)
            print(text)
            result_sheet.cell(row=i+1, column=1, value=url)
            result_sheet.cell(row=i+1, column=2, value=res_code)
            result_sheet.cell(row=i+1, column=3, value=res_time)
            result_sheet.cell(row=i+1, column=4, value=res_size)
            if res_code != 200 and  res_code != 304:
                url_question.put(url)
            else:
                continue
        except:
            url_error.put(url)
            print("异常的链接：%s"%url)


def run_2():
    if norws % 2 == 1:
        for i in range(1,norws//2+2):
            url = sheet.cell((norws+2*(i-1))//2,0).value
            try:
                req = requests.get(url)
                res_code = req.status_code # 请求返回的状态码
                res_time = round(req.elapsed.total_seconds(), 4) # 请求时间，保留4位数
                time_list.append(res_time)
                res_size = round(len(req.text) / 1024, 2)
                text = "进行到第%s条数据,状态码：%s，响应时间：%ss,url：%s，页面大小：%sKB" % (i, res_code, res_time, url, res_size)
                print(text)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=1, value=url)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=2, value=res_code)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=3, value=res_time)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=4, value=res_size)
                if res_code != 200 and  res_code != 304:
                    url_question.put(url)
                else:
                    continue
            except:
                url_error.put(url)
                print("异常的链接：%s" % url)
    else:
        for i in range(1,norws//2+1):
            url = sheet.cell((norws+2*(i-1))//2,0).value
            try:
                req = requests.get(url)
                res_code = req.status_code # 请求返回的状态码
                res_time = round(req.elapsed.total_seconds(), 4) # 请求时间，保留4位数
                time_list.append(res_time)
                res_size = round(len(req.text) / 1024, 2)
                text = "进行到第%s条数据,状态码：%s，响应时间：%ss,url：%s，页面大小：%sKB" % (i, res_code, res_time, url,res_size)
                print(text)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=1, value=url)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=2, value=res_code)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=3, value=res_time)
                result_sheet.cell(row=(norws+2*(i+1))//2, column=4, value=res_size)
                if res_code != 200 and  res_code != 304:
                    url_question.put(url)
                else:
                    continue
            except:
                url_error.put(url)
                print("异常的链接：%s" % url)



if __name__ == '__main__':
    excel_result = Workbook()
    result_sheet = excel_result.create_sheet("Table")
    result_sheet.cell(row=1, column=1, value="url")
    result_sheet.cell(row=1, column=2, value="响应码")
    result_sheet.cell(row=1, column=3, value="响应时间/s")
    result_sheet.cell(row=1, column=4, value="页面大小/KB")
    start_time = time()
    t1 = threading.Thread(target=run_1)
    t2 = threading.Thread(target=run_2)
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    end_time = time()
    run_time = round(end_time - start_time, 2) # 运行时间
    excel_result.save("result.xlsx")
    time_list_sort = time_list.sort() # 把时间列表排序
    final_time_list = time_list[int(len(time_list)*1/10)-1::int(len(time_list)*1/10)] # 把排序后的时间进行切片
    spend_time = "10%耗时：{}s\n20%耗时：{}s\n30%耗时：{}s\n40%耗时：{}s\n50%耗时：{}s\n60%耗时：{}s\n70%耗时：{}s\n80%耗时：{}s\n90%耗时：{}s\n100%耗时：{}s\n".format(final_time_list[0],final_time_list[1], final_time_list[2], final_time_list[3], final_time_list[4], final_time_list[5], final_time_list[6], final_time_list[7], final_time_list[8], final_time_list[9])
    # print("时间列表的长度：%s"%len(time_queue))
    time_list_len = len(time_list) # 时间列表的长度
    url_question_size = url_question.qsize() # 有问题的url的size
    text = "运行的url数量：{}\n有问题的url链接：{}\n有问题的url数量：{}\n运行时间：{}s\n异常的url：{}\n{}".format(time_list_len, url_question.queue, url_question_size, run_time, url_error.queue, spend_time)
    print(text)
    send_dingding(time_list_len, url_question.queue, url_question_size, run_time, url_error.queue, spend_time)

