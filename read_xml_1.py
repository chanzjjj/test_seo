import os
from xml.dom.minidom import parseString
import requests
from openpyxl import load_workbook, Workbook


workbook = Workbook()
ws = workbook.create_sheet("Table")

filepath='C:/Users/mohang/seo-sitemap-temp(1)'
files=os.listdir(filepath)   #获取filepath路径下的所有文件列表
print(files)

path = "C:/Users/mohang/seo-sitemap-temp(1)/"
num = 1
# path2 = "C:/Users/mohang/Downloads/seo-h5-temp/seo-h5-temp/Izmir-restaurants-photos-1.xml"

for i in files:
    path1 = path + i
    print(path1)
    f = open(path1, 'r', encoding="utf-8")
    value = f.read()

    doc = parseString(value)
    collection = doc.documentElement
    returnInfo = collection.getElementsByTagName("url")
    for stu in returnInfo:
        st_loc1 = stu.getAttribute('loc')
        loc = stu.getElementsByTagName('loc')[0].childNodes[0].nodeValue
        ws.cell(row=num, column=1).value = loc
        num += 1
        print(loc)
    f.close()

workbook.save("D:/python_files/python_practice/python_lianxi/test_0203_站点地图.xlsx")

