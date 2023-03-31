import requests
import xlrd
from time import time
import threading
from queue import Queue
from openpyxl import Workbook
from dingding import send_dingding

'''
脚本思路：多线程去运行一份数据，首先把一份数据平均分成若干个相等的子数据，然后一个进程去运行一份子数据
'''


time_list = [] # 存放返回时间的列表
url_question = Queue()  # 有问题的的url队列，非200和304
url_error = Queue()  # 异常的url队列
result_list = [] # 存放返回结果的列表

def get_url(file_name, table_name = "Table", col_num = 0):
    '''
    获取文件数据,参数:
    file_name：文件名
    table_name：表名，默认为Table
    col_num：列数，默认为0
    '''
    url_list = []
    excel = xlrd.open_workbook(file_name)  # 读取excel表
    sheet = excel.sheet_by_name(table_name)  # 读取名为“Table”的sheet
    norws = sheet.nrows  # 总的url数
    for i in range(norws):
        url = sheet.cell(i,col_num).value
        url_list.append(url)
    return url_list, norws

def spilt_data(url_list, num_threading):
    '''把数据平均分成若干份'''
    new_list = []
    n = len(url_list) % num_threading
    n1 = len(url_list) // num_threading
    n2 = n1 + 1
    for i in range(n):
        new_list.append(url_list[n2*i:n2*(i+1)])
    for i in range(n,num_threading):
        new_list.append(url_list[n*n2+(n1*(i-n)):n*n2+(n1*(i-(n-1)))])
    return new_list


def run(*url_list):
    '''运算的主方法，发送请求并获取返回数据，将返回的状态码做判断，非200和304的链接都视为有问题的链接'''
    schedule = 1  # 记录进度
    for url in url_list:
        try:
            req = requests.get(url)
            res_code = req.status_code # 请求返回的状态码
            res_time = round(req.elapsed.total_seconds(),4) # 请求时间，保留4位数
            time_list.append(res_time)
            res_size = round(len(req.text)/1024, 2)
            text = "已检测的页面数：%s,状态码：%s，响应时间：%ss,url：%s，页面大小：%sKB" %(schedule*num_threading, res_code, res_time, url, res_size)
            print(text)
            schedule += 1
            result = {"url": url, "res_code": res_code, "res_time": res_time, "res_size": res_size}
            result_list.append(result)
            if res_code != 200 and  res_code != 304:
                url_question.put(url)
            else:
                continue
        except:
            url_error.put(url)
            print("异常的链接：%s"%url)


def run_threading(data, num_threading):
    '''运行线程'''
    threads = []
    for i in range(num_threading):
        # print(data[i])
        t = threading.Thread(target=run, args=data[i])
        t.start()
        threads.append(t)
    for thread in threads:
        thread.join()

def queue_to_list(data_queue):
    '''队列转列表'''
    len_queue = data_queue.qsize()
    data_list = []
    if len_queue > 0:
        for i in range(len_queue):
            data_list.append(data_queue.get())
        return data_list
    else:
        return []


def sava_to_excel(data, file_name = "result.xlsx"):
    '''保存结果到excel表中'''
    excel_result = Workbook()
    result_sheet = excel_result.create_sheet("Table")
    result_sheet.cell(row=1, column=1, value="url")
    result_sheet.cell(row=1, column=2, value="响应码")
    result_sheet.cell(row=1, column=3, value="响应时间/s")
    result_sheet.cell(row=1, column=4, value="页面大小/KB")
    for index, items in list(enumerate(data)):
        result_sheet.cell(row=index+2, column=1, value=items["url"])
        result_sheet.cell(row=index+2, column=2, value=items["res_code"])
        result_sheet.cell(row=index+2, column=3, value=items["res_time"])
        result_sheet.cell(row=index+2, column=4, value=items["res_size"])
    excel_result.save(file_name)
    print("保存到excel成功")
    excel_result.close()



if __name__ == '__main__':
    url,url_num = get_url("url_files.xlsx")  # 获取数据
    print("需要检测的url数量：%s" % url_num)
    num_threading = 4  # 线程数
    url_list = spilt_data(url, num_threading)
    start_time = time()  # 开始时间
    run_threading(url_list, num_threading)
    end_time = time()  # 结束时间
    sava_to_excel(result_list)  # 保存结果到excel表中
    url_question_list = queue_to_list(url_question)  # 把有问题的url队列转为url列表
    url_error_list = queue_to_list(url_error)  # 把异常的url队列转为url列表
    run_time = round(end_time - start_time, 2)  # 运行时间
    time_list.sort() # 时间列表排序
    spend_time = "10%耗时：{}s\n20%耗时：{}s\n30%耗时：{}s\n40%耗时：{}s\n50%耗时：{}s\n60%耗时：{}s\n70%耗时：{}s\n80%耗时：{}s\n90%耗时：{}s\n95%耗时：{}s\n99%耗时：{}s\n100%耗时：{}s\n".format(time_list[len(time_list)*10//100],time_list[len(time_list)*20//100],time_list[len(time_list)*30//100],time_list[len(time_list)*40//100],time_list[len(time_list)*50//100],time_list[len(time_list)*60//100],time_list[len(time_list)*70//100],time_list[len(time_list)*80//100],time_list[len(time_list)*90//100],time_list[len(time_list)*95//100],time_list[len(time_list)*99//100],time_list[len(time_list)*100//100-1])
    text = "运行的url数量：{}\n有问题的url链接：{}\n有问题的url数量：{}\n运行时间：{}s\n异常的url：{}\n{}".format(len(time_list), url_question_list,
                                                                                     len(url_question_list), run_time,
                                                                                     url_error_list, spend_time)
    print(text)
    send_dingding(url_num, url_question_list, len(url_question_list), run_time, url_error_list, spend_time)
