# -*- coding: UTF-8 -*-
from xml.dom.minidom import parseString
import requests
from openpyxl import load_workbook, Workbook

workbook = Workbook()
ws = workbook.create_sheet("Table")


req = requests.get("https://www.yummyadvisor.com/sitemap.xml")
req.encoding = "utf-8"
res = req.text
# print(res)

doc = parseString(res)
collection = doc.documentElement
returnInfo = collection.getElementsByTagName("sitemap")

num = 1
url_list = []
for stu in returnInfo:
    st_loc = stu.getAttribute('loc')
    url_loc = stu.getElementsByTagName('loc')[0].childNodes[0].nodeValue
    print(url_loc)

    # if "reviews" in url_loc:
    req1 = requests.get(url_loc)
    req1.encoding = "utf-8"
    res1 = req1.text
    doc1 = parseString(res1)
    collection1 = doc1.documentElement
    returnInfo1 = collection1.getElementsByTagName("url")


    for stu1 in returnInfo1:
        st_loc1 = stu1.getAttribute('loc')
        loc1 = stu1.getElementsByTagName('loc')[0].childNodes[0].nodeValue
        # if "menÃ¼ler" in loc1:
        #     loc1 = loc1.replace("menÃ¼ler","menüler")
        # if "fotoÄraflar" in loc1:
        #     loc1 = loc1.replace("fotoÄraflar", "fotoğraflar")
        # if "fotoÄŸraflar" in loc1:
        #     loc1 = loc1.replace("fotoÄŸraflar", "fotoğraflar")
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Balikesir/bkp9pb-Yoruk-Mehmetin-Yeri/menÃ¼ler/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Balikesir/bkp9pb-Yoruk-Mehmetin-Yeri/menüler/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Diyarbakir/jkaox5-Coffee-Yeditepe/menÃ¼ler/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Diyarbakir/jkaox5-Coffee-Yeditepe/menüler/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Manisa/b5cajt-Kapali-Carsi-Manisa-Kebapcisi/menÃ¼ler/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Manisa/b5cajt-Kapali-Carsi-Manisa-Kebapcisi/menüler/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Yozgat/nhhpou-Pizza-House-Coffee-Corner/menÃ¼ler/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Yozgat/nhhpou-Pizza-House-Coffee-Corner/menüler/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Balikesir/bkp9pb-Yoruk-Mehmetin-Yeri/fotoÄŸraflar/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Balikesir/bkp9pb-Yoruk-Mehmetin-Yeri/fotoğraflar/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Diyarbakir/jkaox5-Coffee-Yeditepe/fotoÄŸraflar/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Diyarbakir/jkaox5-Coffee-Yeditepe/fotoğraflar/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Manisa/b5cajt-Kapali-Carsi-Manisa-Kebapcisi/fotoÄŸraflar/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Manisa/b5cajt-Kapali-Carsi-Manisa-Kebapcisi/fotoğraflar/"
        # if loc1 == "https://www.yummyadvisor.com/tr-tr-Yozgat/nhhpou-Pizza-House-Coffee-Corner/fotoÄŸraflar/":
        #     loc1 = "https://www.yummyadvisor.com/tr-tr-Yozgat/nhhpou-Pizza-House-Coffee-Corner/fotoğraflar/"
        ws.cell(row=num, column=1).value = loc1
        num += 1
        print(loc1)

workbook.save("D:/python_files/python_practice/python_lianxi/0223_02.xlsx")






